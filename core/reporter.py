import os
from openpyxl import Workbook
from core.db import ScanDB
from utils.output import log


def export_to_excel(db_path: str, output: str):
    """从 SQLite 导出 Excel"""
    db = ScanDB(db_path)
    wb = Workbook()
    # 移除默认 sheet
    wb.remove(wb.active)

    sheets = {
        "子域名": "SELECT name, domain, ip, cidr, asn, desc, tag, source FROM subdomain",
        "端口服务": "SELECT host, port, service, product, version, title, source FROM port",
        "路径信息": "SELECT url, status, content_length, title FROM path",
        "漏洞信息": "SELECT host, port, service, level, vuln, detail FROM vuln",
        "分析发现": "SELECT category, severity, title, detail, host, port FROM finding",
    }

    for sheet_name, query in sheets.items():
        rows = db.query(query)
        if not rows:
            continue

        ws = wb.create_sheet(sheet_name)
        # 写表头
        headers = list(rows[0].keys())
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        # 写数据
        for row_idx, row in enumerate(rows, 2):
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))

    os.makedirs(os.path.dirname(output), exist_ok=True)
    wb.save(output)
    log.info(f"导出 Excel: {output}")


def export_summary(db_path: str, domain: str) -> str:
    """导出文本摘要"""
    db = ScanDB(db_path)
    scan = db.get_scan(domain)
    if not scan:
        return f"未找到 {domain} 的扫描记录"

    scan_id = scan["id"]
    subdomains = db.query_one("SELECT COUNT(*) as cnt FROM subdomain WHERE scan_id=?", (scan_id,))
    ports = db.query_one("SELECT COUNT(*) as cnt FROM port WHERE scan_id=?", (scan_id,))
    vulns = db.query_one("SELECT COUNT(*) as cnt FROM vuln WHERE scan_id=?", (scan_id,))
    paths = db.query_one("SELECT COUNT(*) as cnt FROM path WHERE scan_id=?", (scan_id,))
    findings = db.query_one("SELECT COUNT(*) as cnt FROM finding WHERE scan_id=?", (scan_id,))

    summary = f"""
=== {domain} 扫描结果 ===
子域名: {subdomains['cnt'] if subdomains else 0}
开放端口: {ports['cnt'] if ports else 0}
路径发现: {paths['cnt'] if paths else 0}
漏洞发现: {vulns['cnt'] if vulns else 0}
分析发现: {findings['cnt'] if findings else 0}

状态: {scan['status']}
创建时间: {scan['created']}
"""
    return summary.strip()
